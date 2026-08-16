"""通用职位表 xlsx 管道测试 — 表头定位/列映射/幂等/缺列置 NULL。

验证 import_position_xlsx.import_xlsx 的核心语义：
- 前 5 行内定位含「职位代码」的表头行（兼容官方表带标题行的结构）
- 表头关键词映射（含括号归一化），缺列置 NULL
- 业务键幂等：重复导入同 xlsx → 0 新增；同 position_code 不同专业行 → 都入库（国考语义）
- 空 position_code 行跳过
"""
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.models.gwy_position import GwyPosition
from app.models.gwy_province_position import GwyProvincePosition
from scripts.import_position_xlsx import import_xlsx, load_mapping

GUOKAO_MAPPING = Path(__file__).resolve().parent.parent / "app" / "crawlers" / "config" / "position_xlsx" / "guokao_2026.yaml"
SHENGKAO_MAPPING = Path(__file__).resolve().parent.parent / "app" / "crawlers" / "config" / "position_xlsx" / "shengkao_gd_2026.yaml"


def _make_guokao_xlsx() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "中央党群机关"
    # 模拟官方结构：标题行在前（不影响表头定位）
    ws.append(["2026年度考试录用公务员招考简章"])
    ws.append(["中央党群机关职位表"])
    ws.append(["部门代码", "部门名称", "用人司局", "招考职位", "职位简介", "职位代码",
               "招考人数", "专业", "学历", "学位", "政治面貌", "工作地点", "备注"])
    ws.append(["001", "中央某部", "办公厅", "一级主任科员", "综合管理", "100110001001",
               1, "计算机类", "本科及以上", "与最高学历相对应的学位", "中共党员", "北京", "需出差"])
    # 同 position_code 第二条（专业/学历不同，国考语义两行都应入库）
    ws.append(["001", "中央某部", "业务司局", "一级主任科员", "业务管理", "100110001001",
               2, "法学类", "仅限硕士研究生", "硕士", "不限", "北京", None])
    # 空 position_code 行应跳过
    ws.append(["001", "中央某部", "办公厅", "临时职位", None, None,
               None, None, None, None, None, None, None])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_shengkao_xlsx() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "县以上机关"
    ws.append(["招考单位", "单位代码", "招考职位", "职位代码", "职位简介", "职位类型",
               "录用人数", "学历", "学位", "本科专业名称及代码", "其他要求", "考区"])
    ws.append(["中共广东省委老干部局", "1990007", "综合岗一级主任科员以下", "19900072641001",
               "从事文字材料工作", "综合管理类职位", 2, "研究生", "硕士以上",
               "政治学(A0302),社会学(A0303)", "中共党员", "广州"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture
def guokao_mapping() -> dict:
    return load_mapping(GUOKAO_MAPPING)


@pytest.fixture
def shengkao_mapping() -> dict:
    return load_mapping(SHENGKAO_MAPPING)


def _write_tmp_xlsx(tmp_path: Path, buf: BytesIO) -> Path:
    p = tmp_path / "positions.xlsx"
    p.write_bytes(buf.getvalue())
    return p


def test_guokao_header_location_and_column_mapping(db_session, tmp_path, guokao_mapping):
    """标题行在前 + 表头关键词映射 + 缺列置 NULL。"""
    xlsx = _write_tmp_xlsx(tmp_path, _make_guokao_xlsx())
    stats = import_xlsx(db_session, xlsx, guokao_mapping)
    assert stats["sheets"] == 1
    assert stats["inserted"] == 2  # 两行有效（空 position_code 跳过）
    assert stats["skipped_bad"] == 1

    rows = db_session.query(GwyPosition).order_by(GwyPosition.recruit_count).all()
    assert len(rows) == 2
    r1 = rows[0]
    assert r1.position_code == "100110001001"
    assert r1.dept_name == "中央某部"
    assert r1.bureau == "办公厅"
    assert r1.position_name == "一级主任科员"
    assert r1.recruit_count == 1
    assert r1.education_req == "本科及以上"
    assert r1.sheet_name == "中央党群机关"
    # 缺列（political_status 未在测试表头中? 已含；检查 interview_ratio 缺列 → NULL）
    assert r1.interview_ratio is None


def test_guokao_same_code_multiple_rows_kept(db_session, tmp_path, guokao_mapping):
    """国考同 position_code 不同专业/学历 → 两行都入库（业务键含 bureau+major+edu+deg）。"""
    xlsx = _write_tmp_xlsx(tmp_path, _make_guokao_xlsx())
    import_xlsx(db_session, xlsx, guokao_mapping)
    codes = [r.bureau for r in db_session.query(GwyPosition).all()]
    assert "办公厅" in codes and "业务司局" in codes
    assert db_session.query(GwyPosition).count() == 2


def test_guokao_idempotent_reimport(db_session, tmp_path, guokao_mapping):
    """重复导入同 xlsx → 0 新增（业务键幂等）。"""
    xlsx = _write_tmp_xlsx(tmp_path, _make_guokao_xlsx())
    import_xlsx(db_session, xlsx, guokao_mapping)
    stats = import_xlsx(db_session, xlsx, guokao_mapping)
    assert stats["inserted"] == 0
    assert stats["skipped_dup"] == 2
    assert db_session.query(GwyPosition).count() == 2


def test_shengkao_mapping_and_duplicate(db_session, tmp_path, shengkao_mapping):
    """省考映射：本科专业列命中 major_req_undergrad + 重复导入幂等。"""
    xlsx = _write_tmp_xlsx(tmp_path, _make_shengkao_xlsx())
    stats = import_xlsx(db_session, xlsx, shengkao_mapping)
    assert stats["inserted"] == 1
    row = db_session.query(GwyProvincePosition).one()
    assert row.province == "广东"
    assert row.dept_name == "中共广东省委老干部局"
    assert row.position_code == "19900072641001"
    assert row.recruit_count == 2
    assert "政治学(A0302)" in (row.major_req_undergrad or "")
    assert row.major_req_grad is None  # 缺列置 NULL

    # 幂等
    stats2 = import_xlsx(db_session, xlsx, shengkao_mapping)
    assert stats2["inserted"] == 0
    assert db_session.query(GwyProvincePosition).count() == 1


def test_header_with_brackets_normalized(db_session, tmp_path, guokao_mapping):
    """表头含括号（如『招考职位(代码)』）也能命中关键词映射。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["部门名称", "招考职位（名称）", "职位代码（唯一）", "招考人数（人）"])
    ws.append(["某部", "一级科员", "A001", 3])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx = _write_tmp_xlsx(tmp_path, buf)
    stats = import_xlsx(db_session, xlsx, guokao_mapping)
    assert stats["inserted"] == 1
    row = db_session.query(GwyPosition).one()
    assert row.position_name == "一级科员"
    assert row.position_code == "A001"
    assert row.recruit_count == 3
