# backend/pipeline/extractors/excel_extractor.py
"""Excel 文本提取器 — 使用 openpyxl。"""
from openpyxl import load_workbook

from app.models.pipeline_enums import ContentType
from pipeline.extractors import ExtractResult

MAX_ROWS = 200


def extract_excel(excel_path: str) -> ExtractResult:
    """逐 sheet 读取 Excel，转为 markdown 表格文本。"""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    parts: list[str] = []
    sheet_names: list[str] = []

    for ws in wb.worksheets:
        sheet_names.append(ws.title)
        rows: list[str] = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= MAX_ROWS:
                break
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append("| " + " | ".join(cells) + " |")
                row_count += 1
        if rows:
            # 表头分隔行
            if len(rows) > 0:
                col_count = rows[0].count("|") - 1
                separator = "| " + " | ".join(["---"] * col_count) + " |"
                parts.append(f"## {ws.title}\n\n{rows[0]}\n{separator}\n" + "\n".join(rows[1:]))

    wb.close()
    full_text = "\n\n".join(parts)
    return ExtractResult(
        text=full_text,
        content_type=ContentType.excel,
        metadata={"sheet_names": sheet_names, "char_count": len(full_text)},
    )
