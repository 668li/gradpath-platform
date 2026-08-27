"""通用官方职位表 xlsx → 职位表入库管道（杠杆化 #2，2026-08-16）。

设计目标：任何省份/年份公务员职位表（官方 Excel）即插即用，告别逐表手写爬虫。
- 列映射声明式 yaml：官方表头关键词 → 模型字段（支持一字段多表头别名）
- 幂等以「业务键内存集合」判断：同 (year, scope, dedup_fields 组合) 已存在 → 跳过，
  不依赖历史 id 算法（历史行哈希不可逆向，旧数据不动）
- 新行主键 id = sha256(dedup_fields 组合)，同组合重复导入天然幂等
- 全 ORM 参数绑定，不拼 SQL；缺列置 NULL，空 position_code 跳过

用法：
  py -3.13 scripts/import_position_xlsx.py <xlsx路径> <mapping.yaml路径>

映射 yaml 格式见 app/crawlers/config/position_xlsx/guokao_2026.yaml。
核心逻辑在 import_xlsx()，可被测试注入内存 DB session。
"""

from __future__ import annotations

import hashlib
import sys
from pathlib import Path
from typing import Any

BACKEND_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_ROOT))

import yaml
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.gwy_position import GwyPosition
from app.models.gwy_province_position import GwyProvincePosition

TABLE_MODELS = {
    "gwy_position": GwyPosition,
    "gwy_province_position": GwyProvincePosition,
}


def _norm_header(text: Any) -> str:
    """表头归一化：去空白、全角→半角、去括号及括号内（如『部门代码(含部门)』）。"""
    s = str(text or "").strip()
    s = s.replace("\u3000", " ").replace("\xa0", " ")
    s = "".join(ch for ch in s if ord(ch) > 32)
    s = s.replace("（", "(").replace("）", ")").replace("：", ":").replace("，", ",")
    # 去括号及内容，避免表头版本差异（如『招考职位（职位简介）』）
    s = "".join(part.split("(")[0] for part in s.split("(")) if "(" in s else s
    return s.replace(" ", "")


def _find_header_row(ws) -> tuple[int, list[str]] | None:
    """定位表头行（前 5 行内包含『职位代码』类列名的行），返回 (行号, 归一化表头列表)。"""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > 5:
            break
        headers = [_norm_header(c) for c in row]
        if any("职位代码" in h or h == "职位代码" for h in headers):
            return row_idx, headers
    return None


def _normalize_cell(value: Any) -> Any:
    """单元格值清洗：None/空白 → None；数字去掉浮点尾。"""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        return s or None
    return value


def load_mapping(yaml_path: Path) -> dict:
    with yaml_path.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def build_column_index(headers: list[str], column_map: dict[str, list[str]]) -> dict[str, int]:
    """官方表头 → 模型字段的列下标映射（表头包含任一关键词即命中）。"""
    index: dict[str, int] = {}
    for field, aliases in column_map.items():
        for alias in aliases:
            key = _norm_header(alias)
            for col_idx, h in enumerate(headers):
                if key and (key in h or h == key):
                    index[field] = col_idx
                    break
            if field in index:
                break
    return index


def build_row_dict(
    row: tuple,
    col_index: dict[str, int],
    column_map: dict[str, list[str]],
    int_fields: list[str],
) -> dict:
    """按列映射把一行 xlsx 转成模型字段 dict（缺列置 None）。"""
    result: dict[str, Any] = {}
    for field in column_map:
        col = col_index.get(field)
        if col is None or col >= len(row):
            result[field] = None
            continue
        value = _normalize_cell(row[col])
        if field in int_fields and value is not None:
            try:
                value = int(str(value).replace(",", "").strip())
            except ValueError:
                value = None
        result[field] = value
    return result


def import_xlsx(db: Session, xlsx_path: Path, mapping: dict) -> dict:
    """核心导入逻辑（可注入测试 DB session）。返回统计 dict。"""
    table = mapping.get("target_table")
    model = TABLE_MODELS.get(table)
    if model is None:
        raise ValueError(f"未知目标表: {table}（支持: {', '.join(TABLE_MODELS)}）")

    year = int(mapping.get("year", 0))
    exam_type = str(mapping.get("exam_type") or "") or None
    province = str(mapping.get("province") or "") or None
    source_url = str(mapping.get("source_url") or "") or None
    dedup_fields = mapping.get("dedup_fields") or ["position_code"]
    int_fields = mapping.get("int_fields") or []
    column_map: dict[str, list[str]] = mapping.get("column_map") or {}

    # 一次性加载既有业务键集合（幂等判断，不依赖历史 id）
    existing_keys: set[tuple] = set()
    for row in db.query(model).all():
        scope_val = exam_type if table == "gwy_position" else province
        parts = [year, scope_val]
        for f in dedup_fields:
            parts.append(getattr(row, f, None))
        existing_keys.add(tuple(parts))

    stats = {"sheets": 0, "read": 0, "inserted": 0, "skipped_dup": 0, "skipped_bad": 0, "errors": 0}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header_info = _find_header_row(ws)
            if header_info is None:
                stats["errors"] += 1
                continue
            _, headers = header_info
            header_row_idx = header_info[0]
            col_index = build_column_index(headers, column_map)
            stats["sheets"] += 1
            sheet_name = ws.title[:50]

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx <= header_row_idx:
                    continue  # 表头行及之前的标题行
                if not any(c is not None and str(c).strip() for c in row):
                    continue
                data = build_row_dict(row, col_index, column_map, int_fields)
                position_code = data.get("position_code")
                if not position_code:
                    stats["skipped_bad"] += 1
                    continue
                scope_val = exam_type if table == "gwy_position" else province
                key_parts = [year, scope_val]
                for f in dedup_fields:
                    key_parts.append(data.get(f))
                if tuple(key_parts) in existing_keys:
                    stats["skipped_dup"] += 1
                    continue
                stats["read"] += 1

                id_source = "|".join(str(k) for k in key_parts if k is not None)
                row_id = hashlib.sha256(id_source.encode()).hexdigest()[:32]
                payload = {
                    **data,
                    "id": row_id,
                    "year": year,
                    "sheet_name": sheet_name,
                    "source_url": source_url,
                }
                if table == "gwy_position":
                    payload["exam_type"] = exam_type
                else:
                    payload["province"] = province
                db.add(model(**payload))
                existing_keys.add(tuple(key_parts))
                stats["inserted"] += 1
                if stats["inserted"] % 500 == 0:
                    db.commit()
        db.commit()
    finally:
        wb.close()
    return stats


def main() -> None:
    if len(sys.argv) < 3:
        print(__doc__)
        return
    xlsx_path = Path(sys.argv[1]).resolve()
    mapping_path = Path(sys.argv[2]).resolve()
    mapping = load_mapping(mapping_path)
    with SessionLocal() as db:
        stats = import_xlsx(db, xlsx_path, mapping)
        model = TABLE_MODELS.get(mapping.get("target_table"))
        total = db.query(model).count()
    print(f"完成: {stats} | {mapping.get('target_table')} 总行数: {total}")


if __name__ == "__main__":
    main()
