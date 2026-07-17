# backend/app/services/export_service.py
"""数据导出服务 — PDF 时间线、JSON 备份、公开技能分享页面。"""
import io
from datetime import date
from uuid import UUID

from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session

from app.models.career_event import CareerEvent
from app.models.community_report import CommunityReport
from app.models.destination_decision import DestinationDecision
from app.models.grad_intel import (
    DarkKnowledge,
    GradAdjustmentInfo,
    GradScorelineRecord,
    GradSchoolIntel,
    GradYanzhaoProgram,
    SelfPositioning,
)
from app.models.interview_report import InterviewReport
from app.models.retrospective import Retrospective
from app.models.skill_node import SkillNode
from app.models.user import User
from app.models.user_setting import UserSetting
from app.services.gamification_service import calculate_xp, get_level


# ======================================================================
# 工具函数
# ======================================================================

def _iso(d) -> str | None:
    """将日期/日期时间转换为 ISO 字符串，None 保持为 None。"""
    if d is None:
        return None
    if hasattr(d, "isoformat"):
        return d.isoformat()
    return str(d)


def _to_str(v) -> str:
    """安全转换为字符串（用于表格单元格）。"""
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return str(v)
    return str(v)


# ======================================================================
# PDF 导出
# ======================================================================

def export_timeline_pdf(db: Session, user_id: UUID) -> bytes:
    """生成 PDF 时间线。

    Sections:
        1. Profile header (name, email)
        2. XP / Level summary
        3. Timeline (decisions + events sorted by date)
        4. Skills summary
        5. Retrospectives list

    使用 reportlab.platypus: SimpleDocTemplate / Paragraph / Spacer / Table,
    通过 io.BytesIO 获取 bytes。
    """
    user = db.query(User).filter(User.id == user_id).first()
    decisions = (
        db.query(DestinationDecision)
        .filter(DestinationDecision.user_id == user_id)
        .order_by(DestinationDecision.decision_date.asc())
        .all()
    )
    events = (
        db.query(CareerEvent)
        .filter(CareerEvent.user_id == user_id)
        .order_by(CareerEvent.event_date.asc())
        .all()
    )
    skills = (
        db.query(SkillNode)
        .filter(SkillNode.user_id == user_id)
        .order_by(SkillNode.category.asc(), SkillNode.name.asc())
        .all()
    )
    retros = (
        db.query(Retrospective)
        .filter(Retrospective.user_id == user_id)
        .order_by(Retrospective.period_end.desc())
        .all()
    )

    xp = calculate_xp(db, user_id)
    level, level_name, current_min, next_min = get_level(xp)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(21 * cm, 29.7 * cm),  # A4
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title="GradPath 时间线",
        author="GradPath",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "GPTitle",
        parent=styles["Title"],
        fontSize=22,
        spaceAfter=6,
        textColor=colors.HexColor("#1e3a8a"),
    )
    subtitle_style = ParagraphStyle(
        "GPSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=12,
    )
    h2_style = ParagraphStyle(
        "GPH2",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceBefore=14,
        spaceAfter=6,
    )
    normal_style = styles["Normal"]

    story: list = []

    # ----- Section 1: Profile header -----
    name = user.name if user else "未知用户"
    email = user.email if user else ""
    story.append(Paragraph("GradPath 成长时间线", title_style))
    profile_lines = [f"<b>姓名：</b>{name}"]
    if email:
        profile_lines.append(f"<b>邮箱：</b>{email}")
    if user and user.school:
        profile_lines.append(f"<b>学校：</b>{user.school}")
    if user and user.major:
        profile_lines.append(f"<b>专业：</b>{user.major}")
    if user and user.graduation_year:
        profile_lines.append(f"<b>毕业年份：</b>{user.graduation_year}")
    story.append(Paragraph("&nbsp;&nbsp;|&nbsp;&nbsp;".join(profile_lines), subtitle_style))
    story.append(Spacer(1, 6))

    # ----- Section 2: XP / Level summary -----
    story.append(Paragraph("游戏化概览", h2_style))
    xp_rows = [
        ["XP", str(xp)],
        ["等级", f"{level} - {level_name}"],
        ["当前等级区间", f"{current_min} - {next_min}"],
        ["去向决策数", str(len(decisions))],
        ["职业事件数", str(len(events))],
        ["技能节点数", str(len(skills))],
        ["复盘数", str(len(retros))],
    ]
    xp_table = Table(xp_rows, colWidths=[5 * cm, 11 * cm])
    xp_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1e3a8a")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbeafe")),
            ]
        )
    )
    story.append(xp_table)

    # ----- Section 3: Timeline (decisions + events sorted by date) -----
    story.append(Paragraph("成长时间线（决策 + 事件）", h2_style))

    timeline_items: list[tuple[date, str, str]] = []
    for d in decisions:
        timeline_items.append(
            (
                d.decision_date,
                f"[决策] {d.destination_type.value}（{d.status.value}）",
                f"信心：{d.confidence}/5；{d.reasoning or ''}".strip("；"),
            )
        )
    for e in events:
        desc = e.title
        if e.description:
            desc += f" — {e.description}"
        timeline_items.append(
            (
                e.event_date,
                f"[事件] {e.event_type.value}",
                desc,
            )
        )
    timeline_items.sort(key=lambda x: x[0])

    if not timeline_items:
        story.append(Paragraph("暂无时间线数据。", normal_style))
    else:
        timeline_rows = [["日期", "类型", "说明"]]
        for d, t, note in timeline_items:
            timeline_rows.append([_iso(d), t, Paragraph(_to_str(note), normal_style)])
        timeline_table = Table(timeline_rows, colWidths=[2.8 * cm, 4.2 * cm, 9 * cm])
        timeline_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.white, colors.HexColor("#f8fafc")]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ]
            )
        )
        story.append(timeline_table)

    # ----- Section 4: Skills summary -----
    story.append(Paragraph("技能树概览", h2_style))
    if not skills:
        story.append(Paragraph("暂无技能节点。", normal_style))
    else:
        skill_rows = [["类别", "名称", "等级", "获得日期", "备注"]]
        for s in skills:
            skill_rows.append(
                [
                    _to_str(s.category),
                    _to_str(s.name),
                    _to_str(s.level),
                    _to_str(_iso(s.acquired_date)),
                    Paragraph(_to_str(s.notes or ""), normal_style),
                ]
            )
        skill_table = Table(
            skill_rows, colWidths=[3 * cm, 4 * cm, 1.5 * cm, 2.8 * cm, 4.7 * cm]
        )
        skill_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (2, 0), (2, -1), "CENTER"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.white, colors.HexColor("#f8fafc")]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ]
            )
        )
        story.append(skill_table)

    # ----- Section 5: Retrospectives list -----
    story.append(Paragraph("阶段复盘", h2_style))
    if not retros:
        story.append(Paragraph("暂无复盘记录。", normal_style))
    else:
        retro_rows = [["周期", "标题", "满意度", "挑战", "收获"]]
        for r in retros:
            retro_rows.append(
                [
                    f"{_iso(r.period_start)} ~ {_iso(r.period_end)}",
                    Paragraph(_to_str(r.title), normal_style),
                    _to_str(r.satisfaction),
                    Paragraph(_to_str(r.challenges or ""), normal_style),
                    Paragraph(_to_str(r.lessons_learned or ""), normal_style),
                ]
            )
        retro_table = Table(
            retro_rows, colWidths=[3.5 * cm, 3 * cm, 1.5 * cm, 4 * cm, 4 * cm]
        )
        retro_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (2, 0), (2, -1), "CENTER"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.white, colors.HexColor("#f8fafc")]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ]
            )
        )
        story.append(retro_table)

    # 页脚生成时间
    story.append(Spacer(1, 20))
    story.append(
        Paragraph(
            f"由 GradPath 自动生成 · {_iso(date.today())}",
            ParagraphStyle(
                "GPFooter",
                parent=normal_style,
                fontSize=8,
                textColor=colors.grey,
                alignment=2,  # right
            ),
        )
    )

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# ======================================================================
# JSON 备份
# ======================================================================

def export_profile_json(db: Session, user_id: UUID) -> dict:
    """导出全部用户数据为 JSON dict。

    结构：
        profile: {name, email, school, major, graduation_year}
        gamification: {xp, level, level_name}
        decisions / events / skills / retrospectives /
        community_reports / interview_reports
    日期序列化为 ISO 字符串，UUID 序列化为字符串。
    """
    user = db.query(User).filter(User.id == user_id).first()

    decisions = (
        db.query(DestinationDecision)
        .filter(DestinationDecision.user_id == user_id)
        .order_by(DestinationDecision.decision_date.asc())
        .all()
    )
    events = (
        db.query(CareerEvent)
        .filter(CareerEvent.user_id == user_id)
        .order_by(CareerEvent.event_date.asc())
        .all()
    )
    skills = (
        db.query(SkillNode)
        .filter(SkillNode.user_id == user_id)
        .all()
    )
    retros = (
        db.query(Retrospective)
        .filter(Retrospective.user_id == user_id)
        .order_by(Retrospective.period_end.desc())
        .all()
    )
    community = (
        db.query(CommunityReport)
        .filter(CommunityReport.user_id == user_id)
        .all()
    )
    interviews = (
        db.query(InterviewReport)
        .filter(InterviewReport.user_id == user_id)
        .all()
    )

    xp = calculate_xp(db, user_id)
    level, level_name, _, _ = get_level(xp)

    profile = {
        "name": user.name if user else None,
        "email": user.email if user else None,
        "school": user.school if user else None,
        "major": user.major if user else None,
        "graduation_year": user.graduation_year if user else None,
    }

    decisions_data = [
        {
            "id": str(d.id),
            "decision_date": _iso(d.decision_date),
            "destination_type": d.destination_type.value if d.destination_type else None,
            "status": d.status.value if d.status else None,
            "details": d.details,
            "reasoning": d.reasoning,
            "confidence": d.confidence,
            "reference_snapshot_id": str(d.reference_snapshot_id) if d.reference_snapshot_id else None,
        }
        for d in decisions
    ]

    events_data = [
        {
            "id": str(e.id),
            "event_date": _iso(e.event_date),
            "event_type": e.event_type.value if e.event_type else None,
            "title": e.title,
            "description": e.description,
            "situation": e.situation,
            "task": e.task,
            "action": e.action,
            "result": e.result,
            "reflection": e.reflection,
            "skills_gained": e.skills_gained,
            "impact_metrics": e.impact_metrics,
            "mood": e.mood,
        }
        for e in events
    ]

    skills_data = [
        {
            "id": str(s.id),
            "name": s.name,
            "category": s.category,
            "level": s.level,
            "parent_id": str(s.parent_id) if s.parent_id else None,
            "acquired_date": _iso(s.acquired_date),
            "notes": s.notes,
        }
        for s in skills
    ]

    retros_data = [
        {
            "id": str(r.id),
            "period_type": r.period_type.value if r.period_type else None,
            "period_start": _iso(r.period_start),
            "period_end": _iso(r.period_end),
            "title": r.title,
            "achievements": r.achievements,
            "challenges": r.challenges,
            "lessons_learned": r.lessons_learned,
            "next_steps": r.next_steps,
            "satisfaction": r.satisfaction,
        }
        for r in retros
    ]

    community_data = [
        {
            "id": str(c.id),
            "school_name": c.school_name,
            "major": c.major,
            "graduation_year": c.graduation_year,
            "degree": c.degree.value if c.degree else None,
            "destination_type": c.destination_type.value if c.destination_type else None,
            "employer": c.employer,
            "city": c.city,
            "industry": c.industry,
            "salary_range": c.salary_range.value if c.salary_range else None,
        }
        for c in community
    ]

    interview_data = [
        {
            "id": str(i.id),
            "company": i.company,
            "position": i.position,
            "city": i.city,
            "interview_year": i.interview_year,
            "rounds": i.rounds,
            "result": i.result.value if i.result else None,
            "dimensions": i.dimensions,
            "difficulty": i.difficulty,
            "summary": i.summary,
        }
        for i in interviews
    ]

    return {
        "profile": profile,
        "gamification": {
            "xp": xp,
            "level": level,
            "level_name": level_name,
        },
        "decisions": decisions_data,
        "events": events_data,
        "skills": skills_data,
        "retrospectives": retros_data,
        "community_reports": community_data,
        "interview_reports": interview_data,
    }


# ======================================================================
# 公开技能分享
# ======================================================================

def get_shareable_skills(db: Session, share_token: str) -> dict | None:
    """根据分享令牌返回公开技能数据。

    - 根据 share_token 查找 UserSetting；未找到或 share_skills_enabled=False 返回 None。
    - 找到后查询该用户姓名与技能树。
    - 返回 {user_name, skills: [...]}，不包含任何其他个人数据。
    """
    if not share_token:
        return None

    setting = (
        db.query(UserSetting)
        .filter(UserSetting.share_token == share_token)
        .first()
    )
    if not setting or not setting.share_skills_enabled:
        return None

    user = db.query(User).filter(User.id == setting.user_id).first()
    if not user:
        return None

    skills = (
        db.query(SkillNode)
        .filter(SkillNode.user_id == setting.user_id)
        .order_by(SkillNode.category.asc(), SkillNode.name.asc())
        .all()
    )

    return {
        "user_name": user.name,
        "skills": [
            {
                "id": str(s.id),
                "name": s.name,
                "category": s.category,
                "level": s.level,
                "parent_id": str(s.parent_id) if s.parent_id else None,
                "acquired_date": _iso(s.acquired_date),
                "notes": s.notes,
            }
            for s in skills
        ],
    }


# ======================================================================
# 考研情报数据导出
# ======================================================================

def export_grad_intel_json(db: Session, user_id: UUID) -> dict:
    """导出当前用户的全部考研情报数据为 JSON dict。

    结构：
        grad_school_intel: 用户收藏的院校情报
        self_positionings: 用户的自我定位记录
        dark_knowledge: 系统预置的暗知识（全局共享）
        yanzhao_programs: 研招网专业目录（全局共享）
        scoreline_records: 复试分数线记录（全局共享）
        adjustment_info: 调剂信息（全局共享）
    """
    grad_intel = (
        db.query(GradSchoolIntel)
        .filter(GradSchoolIntel.user_id == user_id)
        .order_by(GradSchoolIntel.created_at.desc())
        .all()
    )
    positionings = (
        db.query(SelfPositioning)
        .filter(SelfPositioning.user_id == user_id)
        .order_by(SelfPositioning.created_at.desc())
        .all()
    )
    dark_knowledge = (
        db.query(DarkKnowledge)
        .order_by(DarkKnowledge.sort_order.asc())
        .all()
    )
    yanzhao_programs = (
        db.query(GradYanzhaoProgram)
        .order_by(GradYanzhaoProgram.university_name.asc())
        .all()
    )
    scoreline_records = (
        db.query(GradScorelineRecord)
        .order_by(
            GradScorelineRecord.university_name.asc(),
            GradScorelineRecord.year.desc(),
        )
        .all()
    )
    adjustment_info = (
        db.query(GradAdjustmentInfo)
        .order_by(GradAdjustmentInfo.created_at.desc())
        .all()
    )

    return {
        "grad_school_intel": [
            {
                "id": str(g.id),
                "school_name": g.school_name,
                "major_name": g.major_name,
                "school_tier": g.school_tier,
                "year": g.year,
                "background_discrimination": g.background_discrimination,
                "first_choice_protection": g.first_choice_protection,
                "admission_ratio": g.admission_ratio,
                "push_ratio": g.push_ratio,
                "actual_quota": g.actual_quota,
                "score_line": g.score_line,
                "retest_weight": g.retest_weight,
                "retest_format": g.retest_format,
                "score_suppression": g.score_suppression,
                "transfer_friendly": g.transfer_friendly,
                "insider_notes": g.insider_notes,
                "data_sources": g.data_sources,
                "tags": g.tags,
                "ai_summary": g.ai_summary,
                "is_ai_generated": g.is_ai_generated,
                "created_at": _iso(g.created_at),
                "updated_at": _iso(g.updated_at),
            }
            for g in grad_intel
        ],
        "self_positionings": [
            {
                "id": str(s.id),
                "undergrad_tier": s.undergrad_tier,
                "undergrad_major": s.undergrad_major,
                "gpa": s.gpa,
                "gpa_rank": s.gpa_rank,
                "english_level": s.english_level,
                "english_score": s.english_score,
                "research_experience": s.research_experience,
                "competitions": s.competitions,
                "awards": s.awards,
                "internships": s.internships,
                "target_school": s.target_school,
                "target_major": s.target_major,
                "target_region": s.target_region,
                "other_info": s.other_info,
                "ai_assessment": s.ai_assessment,
                "reach_schools": s.reach_schools,
                "target_schools": s.target_schools,
                "safety_schools": s.safety_schools,
                "success_probability": s.success_probability,
                "risk_warnings": s.risk_warnings,
                "created_at": _iso(s.created_at),
                "updated_at": _iso(s.updated_at),
            }
            for s in positionings
        ],
        "dark_knowledge": [
            {
                "id": str(dk.id),
                "stage": dk.stage,
                "category": dk.category,
                "title": dk.title,
                "content": dk.content,
                "importance": dk.importance,
                "common_misconception": dk.common_misconception,
                "actionable_advice": dk.actionable_advice,
                "verification_method": dk.verification_method,
                "tags": dk.tags,
                "sort_order": dk.sort_order,
            }
            for dk in dark_knowledge
        ],
        "yanzhao_programs": [
            {
                "id": str(yp.id),
                "university_name": yp.university_name,
                "department": yp.department,
                "major_name": yp.major_name,
                "degree_type": yp.degree_type,
                "research_directions": yp.research_directions,
                "enrollment_quota": yp.enrollment_quota,
                "tuition": yp.tuition,
                "duration": yp.duration,
                "study_mode": yp.study_mode,
                "admission_requirements": yp.admission_requirements,
                "contact_info": yp.contact_info,
                "source_url": yp.source_url,
                "year": yp.year,
                "data_sources": yp.data_sources,
                "created_at": _iso(yp.created_at),
                "updated_at": _iso(yp.updated_at),
            }
            for yp in yanzhao_programs
        ],
        "scoreline_records": [
            {
                "id": str(sr.id),
                "university_name": sr.university_name,
                "major_name": sr.major_name,
                "degree_type": sr.degree_type,
                "year": sr.year,
                "total_score_line": sr.total_score_line,
                "politics_score": sr.politics_score,
                "foreign_language_score": sr.foreign_language_score,
                "business_1_score": sr.business_1_score,
                "business_2_score": sr.business_2_score,
                "enrollment_count": sr.enrollment_count,
                "application_count": sr.application_count,
                "adjustment_count": sr.adjustment_count,
                "data_sources": sr.data_sources,
                "created_at": _iso(sr.created_at),
                "updated_at": _iso(sr.updated_at),
            }
            for sr in scoreline_records
        ],
        "adjustment_info": [
            {
                "id": str(ai.id),
                "university_name": ai.university_name,
                "department": ai.department,
                "major_name": ai.major_name,
                "degree_type": ai.degree_type,
                "original_major_range": ai.original_major_range,
                "adjustment_quota": ai.adjustment_quota,
                "contact_email": ai.contact_email,
                "contact_phone": ai.contact_phone,
                "deadline": ai.deadline,
                "source_url": ai.source_url,
                "year": ai.year,
                "status": ai.status,
                "data_sources": ai.data_sources,
                "created_at": _iso(ai.created_at),
                "updated_at": _iso(ai.updated_at),
            }
            for ai in adjustment_info
        ],
    }


def export_grad_intel_csv(db: Session, user_id: UUID) -> str:
    """导出当前用户的考研情报数据为 CSV（多节拼接）。

    每个 section 以空行分隔，首行为 section 标题。
    """
    import csv
    import io

    buf = io.StringIO()
    writer = csv.writer(buf)

    # --- Section 1: 院校情报 ---
    writer.writerow(["=== 院校情报 (grad_school_intel) ==="])
    grad_intel = (
        db.query(GradSchoolIntel)
        .filter(GradSchoolIntel.user_id == user_id)
        .order_by(GradSchoolIntel.created_at.desc())
        .all()
    )
    if grad_intel:
        writer.writerow([
            "id", "school_name", "major_name", "school_tier", "year",
            "background_discrimination", "first_choice_protection",
            "admission_ratio", "push_ratio", "actual_quota", "score_line",
            "retest_weight", "score_suppression", "transfer_friendly",
            "insider_notes", "ai_summary", "created_at",
        ])
        for g in grad_intel:
            writer.writerow([
                str(g.id), g.school_name, g.major_name, g.school_tier, g.year,
                g.background_discrimination, g.first_choice_protection,
                g.admission_ratio, g.push_ratio, g.actual_quota, g.score_line,
                g.retest_weight, g.score_suppression, g.transfer_friendly,
                g.insider_notes, g.ai_summary, _iso(g.created_at),
            ])
    writer.writerow([])

    # --- Section 2: 自我定位 ---
    writer.writerow(["=== 自我定位 (self_positionings) ==="])
    positionings = (
        db.query(SelfPositioning)
        .filter(SelfPositioning.user_id == user_id)
        .order_by(SelfPositioning.created_at.desc())
        .all()
    )
    if positionings:
        writer.writerow([
            "id", "undergrad_tier", "undergrad_major", "gpa", "gpa_rank",
            "english_level", "english_score", "target_school", "target_major",
            "target_region", "ai_assessment", "success_probability",
            "created_at",
        ])
        for s in positionings:
            writer.writerow([
                str(s.id), s.undergrad_tier, s.undergrad_major, s.gpa,
                s.gpa_rank, s.english_level, s.english_score, s.target_school,
                s.target_major, s.target_region, s.ai_assessment,
                s.success_probability, _iso(s.created_at),
            ])
    writer.writerow([])

    # --- Section 3: 暗知识 ---
    writer.writerow(["=== 暗知识 (dark_knowledge) ==="])
    dark_knowledge = (
        db.query(DarkKnowledge)
        .order_by(DarkKnowledge.sort_order.asc())
        .all()
    )
    if dark_knowledge:
        writer.writerow([
            "id", "stage", "category", "title", "importance",
            "common_misconception", "actionable_advice", "verification_method",
            "sort_order",
        ])
        for dk in dark_knowledge:
            writer.writerow([
                str(dk.id), dk.stage, dk.category, dk.title, dk.importance,
                dk.common_misconception, dk.actionable_advice,
                dk.verification_method, dk.sort_order,
            ])
    writer.writerow([])

    # --- Section 4: 研招网专业目录 ---
    writer.writerow(["=== 研招网专业目录 (yanzhao_programs) ==="])
    yanzhao_programs = (
        db.query(GradYanzhaoProgram)
        .order_by(GradYanzhaoProgram.university_name.asc())
        .all()
    )
    if yanzhao_programs:
        writer.writerow([
            "id", "university_name", "department", "major_name", "degree_type",
            "enrollment_quota", "tuition", "duration", "study_mode", "year",
            "created_at",
        ])
        for yp in yanzhao_programs:
            writer.writerow([
                str(yp.id), yp.university_name, yp.department, yp.major_name,
                yp.degree_type, yp.enrollment_quota, yp.tuition, yp.duration,
                yp.study_mode, yp.year, _iso(yp.created_at),
            ])
    writer.writerow([])

    # --- Section 5: 分数线记录 ---
    writer.writerow(["=== 分数线记录 (scoreline_records) ==="])
    scoreline_records = (
        db.query(GradScorelineRecord)
        .order_by(
            GradScorelineRecord.university_name.asc(),
            GradScorelineRecord.year.desc(),
        )
        .all()
    )
    if scoreline_records:
        writer.writerow([
            "id", "university_name", "major_name", "degree_type", "year",
            "total_score_line", "politics_score", "foreign_language_score",
            "business_1_score", "business_2_score", "enrollment_count",
            "application_count", "adjustment_count", "created_at",
        ])
        for sr in scoreline_records:
            writer.writerow([
                str(sr.id), sr.university_name, sr.major_name, sr.degree_type,
                sr.year, sr.total_score_line, sr.politics_score,
                sr.foreign_language_score, sr.business_1_score,
                sr.business_2_score, sr.enrollment_count,
                sr.application_count, sr.adjustment_count,
                _iso(sr.created_at),
            ])
    writer.writerow([])

    # --- Section 6: 调剂信息 ---
    writer.writerow(["=== 调剂信息 (adjustment_info) ==="])
    adjustment_info = (
        db.query(GradAdjustmentInfo)
        .order_by(GradAdjustmentInfo.created_at.desc())
        .all()
    )
    if adjustment_info:
        writer.writerow([
            "id", "university_name", "department", "major_name", "degree_type",
            "original_major_range", "adjustment_quota", "contact_email",
            "contact_phone", "deadline", "year", "status", "created_at",
        ])
        for ai in adjustment_info:
            writer.writerow([
                str(ai.id), ai.university_name, ai.department, ai.major_name,
                ai.degree_type, ai.original_major_range, ai.adjustment_quota,
                ai.contact_email, ai.contact_phone, ai.deadline, ai.year,
                ai.status, _iso(ai.created_at),
            ])

    return buf.getvalue()


# ======================================================================
# 考研情报 PDF 导出
# ======================================================================

def export_grad_intel_pdf(db: Session, user_id: UUID) -> bytes:
    """生成考研情报报告 PDF。

    包含：院校情报汇总、自我定位、暗知识精选。
    """
    from datetime import date as _date

    user = db.query(User).filter(User.id == user_id).first()
    grad_intel = (
        db.query(GradSchoolIntel)
        .filter(GradSchoolIntel.user_id == user_id)
        .order_by(GradSchoolIntel.school_tier, GradSchoolIntel.school_name)
        .all()
    )
    positionings = (
        db.query(SelfPositioning)
        .filter(SelfPositioning.user_id == user_id)
        .order_by(SelfPositioning.created_at.desc())
        .all()
    )
    dark_knowledge = (
        db.query(DarkKnowledge)
        .order_by(DarkKnowledge.importance.desc(), DarkKnowledge.sort_order.asc())
        .limit(30)
        .all()
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(21 * cm, 29.7 * cm),
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title="GradPath 考研情报报告",
        author="GradPath",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "GIPdfTitle",
        parent=styles["Title"],
        fontSize=22,
        spaceAfter=6,
        textColor=colors.HexColor("#1e3a8a"),
    )
    subtitle_style = ParagraphStyle(
        "GIPdfSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=12,
    )
    h2_style = ParagraphStyle(
        "GIPdfH2",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceBefore=14,
        spaceAfter=6,
    )
    normal_style = styles["Normal"]

    story: list = []

    # ----- Header -----
    name = user.name if user else "未知用户"
    story.append(Paragraph("GradPath 考研情报报告", title_style))
    story.append(Paragraph(f"用户：{name}　|　生成日期：{_iso(_date.today())}", subtitle_style))
    story.append(Spacer(1, 6))

    # ----- Section 1: 院校情报 -----
    story.append(Paragraph("院校情报汇总", h2_style))
    if not grad_intel:
        story.append(Paragraph("暂无院校情报。", normal_style))
    else:
        intel_rows = [["院校", "专业", "层次", "分数线", "报录比", "保护一志愿", "压分"]]
        for g in grad_intel:
            intel_rows.append([
                _to_str(g.school_name),
                _to_str(g.major_name),
                _to_str(g.school_tier),
                _to_str(g.score_line),
                _to_str(g.admission_ratio),
                _to_str(g.first_choice_protection),
                _to_str(g.score_suppression),
            ])
        intel_table = Table(intel_rows, colWidths=[2.8*cm, 2.5*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm])
        intel_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (3, 0), (3, -1), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ]))
        story.append(intel_table)

        # 内部备注
        for g in grad_intel:
            if g.insider_notes:
                story.append(Spacer(1, 4))
                story.append(Paragraph(
                    f"<b>{g.school_name} - {g.major_name} 内部备注：</b>{_to_str(g.insider_notes)}",
                    normal_style,
                ))

    # ----- Section 2: 自我定位 -----
    story.append(Paragraph("自我定位与推荐", h2_style))
    if not positionings:
        story.append(Paragraph("暂无自我定位记录。", normal_style))
    else:
        for s in positionings[:3]:
            story.append(Paragraph(
                f"<b>目标：</b>{_to_str(s.target_school)} - {_to_str(s.target_major)}",
                normal_style,
            ))
            if s.ai_assessment:
                story.append(Paragraph(f"AI 评估：{_to_str(s.ai_assessment)}", normal_style))
            if s.reach_schools:
                story.append(Paragraph(f"冲刺院校：{_to_str(s.reach_schools)}", normal_style))
            if s.target_schools:
                story.append(Paragraph(f"稳妥院校：{_to_str(s.target_schools)}", normal_style))
            if s.safety_schools:
                story.append(Paragraph(f"保底院校：{_to_str(s.safety_schools)}", normal_style))
            if s.success_probability:
                story.append(Paragraph(f"成功概率：{_to_str(s.success_probability)}", normal_style))
            story.append(Spacer(1, 8))

    # ----- Section 3: 暗知识精选 -----
    story.append(Paragraph("暗知识精选（重要度排序）", h2_style))
    if not dark_knowledge:
        story.append(Paragraph("暂无暗知识数据。", normal_style))
    else:
        dk_rows = [["标题", "阶段", "分类", "重要度", "常见误区", "行动建议"]]
        for dk in dark_knowledge:
            dk_rows.append([
                Paragraph(_to_str(dk.title), normal_style),
                _to_str(dk.stage),
                _to_str(dk.category),
                _to_str(dk.importance),
                Paragraph(_to_str(dk.common_misconception or ""), normal_style),
                Paragraph(_to_str(dk.actionable_advice or ""), normal_style),
            ])
        dk_table = Table(dk_rows, colWidths=[2.5*cm, 1.5*cm, 1.5*cm, 1.2*cm, 3.5*cm, 3.5*cm])
        dk_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ]))
        story.append(dk_table)

    # 页脚
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"由 GradPath 自动生成 · {_iso(_date.today())}",
        ParagraphStyle("GIPdfFooter", parent=normal_style, fontSize=8, textColor=colors.grey, alignment=2),
    ))

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# ======================================================================
# 暗知识按阶段批量导出 PDF
# ======================================================================

def export_dark_knowledge_by_stage_pdf(db: Session, stage: str | None = None) -> bytes:
    """按阶段导出暗知识为 PDF。

    stage 为 None 时导出全部。
    """
    from datetime import date as _date

    query = db.query(DarkKnowledge)
    if stage:
        query = query.filter(DarkKnowledge.stage == stage)
    items = query.order_by(DarkKnowledge.stage, DarkKnowledge.sort_order.asc()).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(21 * cm, 29.7 * cm),
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title="GradPath 暗知识导出",
        author="GradPath",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("DKPdfTitle", parent=styles["Title"], fontSize=22, spaceAfter=6, textColor=colors.HexColor("#1e3a8a"))
    h2_style = ParagraphStyle("DKPdfH2", parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#1e3a8a"), spaceBefore=12, spaceAfter=4)
    normal_style = styles["Normal"]

    story: list = []

    stage_label = f"（{stage} 阶段）" if stage else "（全部阶段）"
    story.append(Paragraph(f"暗知识导出 {stage_label}", title_style))
    story.append(Paragraph(f"共 {len(items)} 条　|　生成日期：{_iso(_date.today())}", ParagraphStyle("DKPdfSub", parent=normal_style, fontSize=10, textColor=colors.grey, spaceAfter=12)))

    # 按阶段分组
    grouped: dict[str, list] = {}
    for dk in items:
        grouped.setdefault(dk.stage or "未分类", []).append(dk)

    for stage_name, stage_items in grouped.items():
        story.append(Paragraph(f"{stage_name}（{len(stage_items)} 条）", h2_style))
        rows = [["标题", "分类", "重要度", "常见误区", "行动建议", "验证方式"]]
        for dk in stage_items:
            rows.append([
                Paragraph(_to_str(dk.title), normal_style),
                _to_str(dk.category),
                _to_str(dk.importance),
                Paragraph(_to_str(dk.common_misconception or ""), normal_style),
                Paragraph(_to_str(dk.actionable_advice or ""), normal_style),
                Paragraph(_to_str(dk.verification_method or ""), normal_style),
            ])
        tbl = Table(rows, colWidths=[2.5*cm, 1.2*cm, 1*cm, 3*cm, 3*cm, 3*cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ]))
        story.append(tbl)

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"由 GradPath 自动生成 · {_iso(_date.today())}",
        ParagraphStyle("DKPdfFooter", parent=normal_style, fontSize=8, textColor=colors.grey, alignment=2),
    ))

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

# ======================================================================
# 考研情报 Excel 导出
# ======================================================================

def export_grad_intel_excel(db: Session, user_id: UUID) -> bytes:
    """导出当前用户的考研情报数据为 Excel (.xlsx)。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D2D6DC'),
        right=Side(style='thin', color='D2D6DC'),
        top=Side(style='thin', color='D2D6DC'),
        bottom=Side(style='thin', color='D2D6DC'),
    )
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    def _write_headers(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.freeze_panes = 'A2'

    def _write_data(ws, data, start_row=2):
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=start_row + row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = alt_fill

    def _auto_width(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[col_letter].width = adjusted_width

    # Sheet 1: 院校情报
    ws1 = wb.active
    ws1.title = '院校情报'
    # 优化：添加 limit 防止全量查询导致 OOM
    grad_intel = db.query(GradSchoolIntel).filter(GradSchoolIntel.user_id == user_id).order_by(GradSchoolIntel.school_tier, GradSchoolIntel.school_name).limit(500).all()
    headers1 = ['院校', '专业', '层次', '年份', '出身歧视', '保护一志愿', '报录比', '推免比例', '实际招生', '分数线', '复试权重', '复试形式', '压分', '调剂友好', '内部备注', '数据来源', '标签', 'AI摘要', '创建时间']
    _write_headers(ws1, headers1)
    data1 = [[g.school_name, g.major_name, g.school_tier, g.year, g.background_discrimination, g.first_choice_protection, g.admission_ratio, g.push_ratio, g.actual_quota, g.score_line, g.retest_weight, g.retest_format, g.score_suppression, g.transfer_friendly, g.insider_notes, g.data_sources, g.tags, g.ai_summary, _iso(g.created_at)] for g in grad_intel]
    _write_data(ws1, data1)
    _auto_width(ws1)

    # Sheet 2: 自我定位
    ws2 = wb.create_sheet('自我定位')
    positionings = db.query(SelfPositioning).filter(SelfPositioning.user_id == user_id).order_by(SelfPositioning.created_at.desc()).limit(100).all()
    headers2 = ['本科层次', '本科专业', 'GPA', 'GPA排名', '英语水平', '英语成绩', '科研经历', '竞赛', '获奖', '实习', '目标院校', '目标专业', '目标地区', '其他信息', 'AI评估', '冲刺院校', '稳妥院校', '保底院校', '成功概率', '风险警告', '创建时间']
    _write_headers(ws2, headers2)
    data2 = [[s.undergrad_tier, s.undergrad_major, s.gpa, s.gpa_rank, s.english_level, s.english_score, s.research_experience, s.competitions, s.awards, s.internships, s.target_school, s.target_major, s.target_region, s.other_info, s.ai_assessment, s.reach_schools, s.target_schools, s.safety_schools, s.success_probability, s.risk_warnings, _iso(s.created_at)] for s in positionings]
    _write_data(ws2, data2)
    _auto_width(ws2)

    # Sheet 3: 暗知识
    ws3 = wb.create_sheet('暗知识')
    dark_knowledge = db.query(DarkKnowledge).order_by(DarkKnowledge.stage, DarkKnowledge.sort_order).limit(1000).all()
    headers3 = ['阶段', '分类', '标题', '内容', '重要度', '常见误区', '行动建议', '验证方式', '标签', '排序']
    _write_headers(ws3, headers3)
    data3 = [[dk.stage, dk.category, dk.title, dk.content, dk.importance, dk.common_misconception, dk.actionable_advice, dk.verification_method, dk.tags, dk.sort_order] for dk in dark_knowledge]
    _write_data(ws3, data3)
    _auto_width(ws3)

    # Sheet 4: 研招网专业目录
    ws4 = wb.create_sheet('研招网专业目录')
    yanzhao_programs = db.query(GradYanzhaoProgram).order_by(GradYanzhaoProgram.university_name).limit(5000).all()
    headers4 = ['院校', '院系', '专业', '学位类型', '研究方向', '招生计划', '学费', '学制', '学习方式', '报考要求', '联系方式', '来源链接', '年份', '数据来源', '创建时间']
    _write_headers(ws4, headers4)
    data4 = [[yp.university_name, yp.department, yp.major_name, yp.degree_type, yp.research_directions, yp.enrollment_quota, yp.tuition, yp.duration, yp.study_mode, yp.admission_requirements, yp.contact_info, yp.source_url, yp.year, yp.data_sources, _iso(yp.created_at)] for yp in yanzhao_programs]
    _write_data(ws4, data4)
    _auto_width(ws4)

    # Sheet 5: 分数线记录
    ws5 = wb.create_sheet('分数线记录')
    scoreline_records = db.query(GradScorelineRecord).order_by(GradScorelineRecord.university_name, GradScorelineRecord.year.desc()).limit(5000).all()
    headers5 = ['院校', '专业', '学位类型', '年份', '总分线', '政治', '外语', '业务课一', '业务课二', '招生人数', '报考人数', '调剂人数', '数据来源', '创建时间']
    _write_headers(ws5, headers5)
    data5 = [[sr.university_name, sr.major_name, sr.degree_type, sr.year, sr.total_score_line, sr.politics_score, sr.foreign_language_score, sr.business_1_score, sr.business_2_score, sr.enrollment_count, sr.application_count, sr.adjustment_count, sr.data_sources, _iso(sr.created_at)] for sr in scoreline_records]
    _write_data(ws5, data5)
    _auto_width(ws5)

    # Sheet 6: 调剂信息
    ws6 = wb.create_sheet('调剂信息')
    adjustment_info = db.query(GradAdjustmentInfo).order_by(GradAdjustmentInfo.created_at.desc()).limit(2000).all()
    headers6 = ['院校', '院系', '专业', '学位类型', '原专业范围', '调剂名额', '联系邮箱', '联系电话', '截止日期', '来源链接', '年份', '状态', '数据来源', '创建时间']
    _write_headers(ws6, headers6)
    data6 = [[ai.university_name, ai.department, ai.major_name, ai.degree_type, ai.original_major_range, ai.adjustment_quota, ai.contact_email, ai.contact_phone, ai.deadline, ai.source_url, ai.year, ai.status, ai.data_sources, _iso(ai.created_at)] for ai in adjustment_info]
    _write_data(ws6, data6)
    _auto_width(ws6)

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
